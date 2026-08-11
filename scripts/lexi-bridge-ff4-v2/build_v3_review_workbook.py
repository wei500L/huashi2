#!/usr/bin/env python3
"""Build a reviewer-friendly XLSX directly from the V3 research seed."""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
OUT = ROOT / "docs" / "data" / "lexi-bridge-ff4-v2"
SEED = ROOT / "app-server" / "src" / "main" / "resources" / "assessment-seeds" / "LEXIBRIDGE_RESEARCH_V3.json"
TARGET = OUT / "法语专四假朋友题库_V3.xlsx"

TITLE_FILL = PatternFill("solid", fgColor="17365D")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBTITLE_FILL = PatternFill("solid", fgColor="D9EAF7")
WHITE_BOLD = Font(color="FFFFFF", bold=True, size=14)
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")

SECTION_SHEETS = {
    "FF4_WORD_MEANING": ("单选", "题型一｜正确中文义 + 英文迁移义 + 两个同词性干扰项。"),
    "FF4_SENTENCE_SELECTION": ("选词填空", "题型二｜使用可定位 TEM4 原句，选择语境中的正确近义表达。"),
    "FF4_TRUE_FALSE": ("判断", "题型三｜严格按 0811 文档使用英文迁移义，全部答案自然为 F。"),
    "FF4_SPELLING": ("拼写", "题型四｜根据中文释义拼写法语词；首次错误后仅提示首字母。"),
}


def load_jsonl(path: Path) -> list[dict]:
    return [json.loads(line) for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]


def setup_sheet(sheet, title: str, subtitle: str, columns: list[str]) -> None:
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    title_cell = sheet.cell(1, 1, title)
    title_cell.fill = TITLE_FILL
    title_cell.font = WHITE_BOLD
    title_cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 26
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
    subtitle_cell = sheet.cell(2, 1, subtitle)
    subtitle_cell.fill = SUBTITLE_FILL
    subtitle_cell.alignment = WRAP
    sheet.row_dimensions[2].height = 36
    sheet.append(columns)
    for cell in sheet[3]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP
    sheet.freeze_panes = "A4"
    sheet.auto_filter.ref = f"A3:{get_column_letter(len(columns))}3"


def option_map(item: dict) -> dict[str, str]:
    return {option["key"]: option["label"] for option in item.get("options", [])}


def write_question_sheets(workbook: Workbook, seed: dict, records: dict[str, dict]) -> None:
    columns = [
        "题目编号", "词条ID", "目标词", "题干", "选项A/V", "选项B/F", "选项C", "选项D",
        "标准答案", "词义辨析解析", "来源页码", "词汇审查状态", "教研审查状态",
    ]
    for section_code, (sheet_name, subtitle) in SECTION_SHEETS.items():
        sheet = workbook.create_sheet(sheet_name)
        setup_sheet(sheet, f"法语专四假朋友题库 V3｜{sheet_name}", subtitle, columns)
        items = [item for item in seed["items"] if item["sectionCode"] == section_code]
        for item in items:
            record = records.get(item.get("targetWord"), {})
            options = option_map(item)
            if section_code == "FF4_TRUE_FALSE":
                option_a = "Vrai（正确）"
                option_b = "Faux（错误）"
            else:
                option_a = options.get("A")
                option_b = options.get("B")
            pages = f"TEM4 p.{item.get('tem4PdfPage')}；假朋友词典 p.{item.get('falseFriendsPdfPage')}"
            sheet.append([
                item["itemCode"], record.get("wordId"), item.get("targetWord"), item.get("stemText"),
                option_a, option_b, options.get("C"), options.get("D"),
                "；".join(item.get("correctAnswers") or []), item.get("explanationText"), pages,
                item.get("lexicalReviewStatus"), item.get("pedagogicReviewStatus"),
            ])
        for row in sheet.iter_rows(min_row=4):
            for cell in row:
                cell.alignment = WRAP
        widths = [18, 24, 18, 52, 28, 28, 28, 28, 12, 72, 30, 24, 24]
        for index, width in enumerate(widths, 1):
            sheet.column_dimensions[get_column_letter(index)].width = width


def write_overview(workbook: Workbook, seed: dict) -> None:
    sheet = workbook.active
    sheet.title = "V3说明"
    scored = [item for item in seed["items"] if item.get("scored")]
    counts = {code: sum(1 for item in scored if item["sectionCode"] == code) for code in SECTION_SHEETS}
    rows = [
        ("问卷代码", seed["packageCode"]),
        ("问卷标题", seed["questionnaire"]["title"]),
        ("问卷状态", seed["questionnaire"]["status"]),
        ("计分版本", seed["questionnaire"]["scoringVersion"]),
        ("计分题总数", len(scored)),
        ("题型一", counts["FF4_WORD_MEANING"]),
        ("题型二", counts["FF4_SENTENCE_SELECTION"]),
        ("题型三", counts["FF4_TRUE_FALSE"]),
        ("题型四", counts["FF4_SPELLING"]),
        ("内容状态", "APPROVED（全部题目已通过内容与结构审计）"),
        ("覆盖状态", "MAXIMUM_RULE_COMPLIANT_COVERAGE（与完整生产题库逐题一致）"),
        ("导入状态", "READY_FOR_IMPORT_VALIDATION"),
        ("发布状态", "NOT_DEPLOYED（未连接生产库、未创建 release）"),
        ("题型三说明", "严格执行 0811 文档模板，答案全部为 F，不另造控制题。"),
    ]
    sheet.append(["法语专四假朋友题库 V3｜生成说明", None])
    sheet.merge_cells("A1:B1")
    sheet["A1"].fill = TITLE_FILL
    sheet["A1"].font = WHITE_BOLD
    sheet.append(["字段", "内容"])
    for cell in sheet[2]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for row in rows:
        sheet.append(row)
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 90
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = WRAP
    sheet.freeze_panes = "A3"


def write_basic_info(workbook: Workbook, seed: dict) -> None:
    sheet = workbook.create_sheet("基本信息")
    columns = ["题目编号", "题型", "题目", "是否必填", "选项", "显示条件"]
    setup_sheet(sheet, "Lexi-Bridge V3｜基本信息", "本页完整对应 V3 种子中的 BASIC_INFO 分区。", columns)
    for item in [item for item in seed["items"] if item["sectionCode"] == "BASIC_INFO"]:
        options = "；".join(f"{option['key']}={option['label']}" for option in item.get("options", []))
        condition = json.dumps(item.get("displayCondition"), ensure_ascii=False) if item.get("displayCondition") else None
        sheet.append([item["itemCode"], item["questionType"], item["stemText"],
                      "是" if item.get("requiredAnswer") else "否", options, condition])
    for row in sheet.iter_rows(min_row=4):
        for cell in row:
            cell.alignment = WRAP
    for index, width in enumerate([24, 24, 90, 12, 70, 45], 1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def write_review_issues(workbook: Workbook, seed: dict) -> None:
    sheet = workbook.create_sheet("审查问题")
    columns = ["问题代码", "严重级别", "题目编号", "说明"]
    setup_sheet(sheet, "Lexi-Bridge V3｜审查问题", "生产审查已完成；本页保留生成器输出的问题记录。", columns)
    for issue in seed.get("reviewIssues", []):
        sheet.append([issue.get("issueCode"), issue.get("severity"), issue.get("itemCode"), issue.get("description")])
    for row in sheet.iter_rows(min_row=4):
        for cell in row:
            cell.alignment = WRAP
    for index, width in enumerate([30, 22, 22, 90], 1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def main() -> None:
    seed = json.loads(SEED.read_text(encoding="utf-8"))
    records = {record["frenchWord"]: record for record in load_jsonl(OUT / "candidate-adjudication.jsonl")}
    workbook = Workbook()
    write_overview(workbook, seed)
    write_question_sheets(workbook, seed, records)
    write_basic_info(workbook, seed)
    write_review_issues(workbook, seed)
    workbook.save(TARGET)
    print(TARGET)


if __name__ == "__main__":
    main()
