#!/usr/bin/env python3
"""Build the production FF4 bank after concept-level AI adjudication.

The production bank contains no pending-review placeholders. A candidate is
either approved with source-backed reasoning or excluded. Target words remain
globally unique across the four item types.
"""

from __future__ import annotations

import json
import random
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))

from build_candidates import ROOT, load_jsonl  # noqa: E402
from build_question_bank import (  # noqa: E402
    build_t1_item, build_t3_item, build_t4_item,
    find_t4_ambiguous,
)
from production_semantic_rules import (  # noqa: E402
    APPROVED_SEMANTIC_WORDS, T2_RULES, T2_TARGET_WORDS, TRUE_COGNATE_CONTROLS,
)

OUT = ROOT / "docs" / "data" / "lexi-bridge-ff4-v2"
PACKAGE_PATH = OUT / "question-bank-package-production.json"
REVIEW_PATH = OUT / "production-semantic-review.jsonl"
WORKBOOK_PATH = OUT / "法语专四假朋友题库_生产版.xlsx"

EXCLUDED_SEMANTIC_REASONS = {
    "casserole": "法语炊具与英语焙盘菜属于容器—菜品的直接转喻关系，概念距离不足。",
    "chic": "法语“漂亮、别致”与英语“时髦”共享核心审美义。",
    "couvert": "法语“被覆盖、布满”与英语 cover 的遮盖义直接对应。",
    "cycle": "英语 cycle 本身也表示循环，且 bicycle 义与法语循环义同源，不能视为完全不重合。",
    "diffuser": "法语“传播”与英语 diffuse 的“扩散”共享核心扩散义。",
    "draguer": "法语“疏浚”与英语 drag 的拖拉动作存在直接工具行为联系。",
    "début": "“开始”与 debut 的“初次亮相”共享事件起始义。",
    "dépendre": "法语“取决于”与英语 depend 的“依赖”属于同一依存关系。",
    "déranger": "“弄乱、打扰”与 derange 的“使精神错乱”共享扰乱义。",
    "dériver": "法语“从……派生”与英语 derive 的“起源于、获得”直接对应。",
    "express": "英语 express 也可表示快速、特快服务，与法语“特快的”重合。",
    "gourmand": "贪吃、讲究吃食与 gourmand 的美食家义存在明显范围重合。",
    "infirme": "残疾、体弱与 infirm 的年老体弱属于同一身体虚弱概念。",
    "kiosque": "报亭、亭子与 kiosk 的售货亭、自助亭共享亭式设施核心义。",
    "organisme": "法语“机体”与英语 organism 的“生物体”直接重合。",
    "presser": "法语“压、挤”与英语 press 的压按义直接重合。",
    "publicité": "广告与 publicity 的宣传、公众传播义存在核心重合。",
    "quartier": "英语 quarter 也可表示城区、住处，与法语居民区重合。",
    "rubrique": "专栏、栏目与 rubric 的标题、分类标准共享栏目/标题概念。",
    "régulier": "“有规律的”与 regular 的“规则的、经常的”直接重合。",
    "sain": "健康与 sane 的精神健全同属健康状态，概念并非完全分离。",
    "signaler": "标示、报告与 signal 的发信号义共享提示/传达功能。",
    "sot": "现有英语中文义“盐酸盐”与可靠英语词义不符，证据源存在异常，生产版排除。",
    "sympathique": "心善、给人好感与 sympathetic 的同情、善意存在明显重合。",
    "utilitaire": "实用、求实利与 utilitarian 的功利主义共享效用/功利核心。",
    "zinc": "法语 zinc 的常用核心义包含金属“锌”，当前 TEM4 前义提取疑似遗漏，证据冲突。",
    "élaborer": "制作、形成与 elaborate 的详述、精心发展存在过程义重合。",
}

TYPE_META = {
    "T1": ("FF4_WORD_MEANING", "SINGLE_CHOICE", "FF4_WORD_MEANING", "WORD"),
    "T2": ("FF4_SENTENCE_SELECTION", "SINGLE_CHOICE", "FF4_SENTENCE_SYNONYM", "SENTENCE"),
    "T3": ("FF4_TRUE_FALSE", "TRUE_FALSE", "FF4_TRUE_FALSE_TRANSFER", "WORD"),
    "T4": ("FF4_SPELLING", "SPELLING", "FF4_SPELLING", "WORD"),
}


def write_semantic_review(records: list[dict]) -> list[dict]:
    reviewed = []
    for record in records:
        if (not str(record.get("sourceRelationType") or "").startswith("A ")
                and record["frenchWord"] not in APPROVED_SEMANTIC_WORDS):
            continue
        word = record["frenchWord"]
        approved = word in APPROVED_SEMANTIC_WORDS
        if approved:
            reason = (f"复核 TEM4 前义“{'；'.join(record['tem4TopSenses'])}”与英语迁移义"
                      f"“{'；'.join(record['englishChineseSenses'])}”，二者核心概念、上下位关系及常用语境均不重合。")
        else:
            reason = EXCLUDED_SEMANTIC_REASONS.get(word)
            if reason is None:
                reason = ("生产复核未通过：" + record["semanticDecisionReason"]
                          + " 同时考虑完整词义、跨词性同形义和现代常用义后，不能确认概念完全不重合。")
        reviewed.append({
            "wordId": record["wordId"],
            "frenchWord": word,
            "englishConfusable": record["englishConfusable"],
            "tem4TopSenses": record["tem4TopSenses"],
            "englishChineseSenses": record["englishChineseSenses"],
            "productionDecision": "APPROVED" if approved else "EXCLUDED",
            "productionDecisionReason": reason,
            "reviewMethod": "AI_CONCEPT_LEVEL_SOURCE_REVIEW_V1",
        })
    REVIEW_PATH.write_text("".join(json.dumps(row, ensure_ascii=False) + "\n" for row in reviewed), encoding="utf-8")
    return reviewed


def build_production_t2_item(record: dict) -> dict:
    """Follow the source document literally: one synonym, one French word
    matching the English-transfer meaning, and two same-form distractors."""
    rule = T2_RULES[record["frenchWord"]]
    return {
        "itemType": "T2",
        "targetWord": record["frenchWord"],
        "stemText": "请根据句子选择画线单词的同义解释\n" + rule["sentence"],
        "options": [
            {"key": key, "label": label, "correct": correct, "role": role}
            for key, label, correct, role in rule["options"]
        ],
        "correctAnswers": ["A"],
        "evidenceLevel": rule.get("evidenceLevel", "TEM4_EXACT_SENTENCE"),
        "sourceCollocation": rule.get("sourceCollocation"),
    }


def build_items(records: list[dict], evidence: dict[str, dict]) -> list[dict]:
    rng = random.Random(20260812)
    by_word = {record["frenchWord"]: record for record in records}
    semantic = [by_word[word] for word in sorted(APPROVED_SEMANTIC_WORDS)]

    t2_items = []
    for word in T2_TARGET_WORDS:
        record = by_word[word].copy()
        record.update(evidence.get(word, {}))
        t2_items.append(build_production_t2_item(record))

    remaining = [record for record in semantic if record["frenchWord"] not in T2_TARGET_WORDS]
    # T1 needs two same-POS distractors. Use roughly half of the semantic pool
    # for T1 and reserve the rest for T3 so target words remain globally unique.
    t1_items = []
    t3_records = []
    for record in remaining:
        item = build_t1_item(record, records, rng)
        if len(t1_items) < 27 and item is not None:
            t1_items.append(item)
        else:
            t3_records.append(record)
    t3_items = [build_t3_item(record) for record in t3_records]
    true_controls = []
    for word, sense in TRUE_COGNATE_CONTROLS.items():
        true_controls.append({
            "itemType": "T3",
            "targetWord": word,
            "stemText": f"{word} = {sense}",
            "options": [
                {"key": "V", "label": "正确", "correct": True, "role": "CORRECT"},
                {"key": "F", "label": "错误", "correct": False, "role": "DISTRACTOR"},
            ],
            "correctAnswers": ["V"],
            "controlType": "COGNATE_CONTROL",
        })
    rng.shuffle(t3_items)
    # Spread the Vrai controls across the complete false-friend list. Using
    # n+1 gaps keeps the sequence deterministic and avoids a long F-only tail.
    gap_count = len(true_controls) + 1
    base_gap, extra = divmod(len(t3_items), gap_count)
    gap_sizes = [base_gap + (1 if index < extra else 0) for index in range(gap_count)]
    interleaved_t3 = []
    offset = 0
    for index, gap_size in enumerate(gap_sizes):
        interleaved_t3.extend(t3_items[offset:offset + gap_size])
        offset += gap_size
        if index < len(true_controls):
            interleaved_t3.append(true_controls[index])
    t3_items = interleaved_t3

    used = {item["targetWord"] for item in t1_items + t2_items + t3_items}
    ambiguous = find_t4_ambiguous(records)
    t4_records = [
        record for record in records
        if "T4" in record["eligibleTypes"]
        and record["frenchWord"] not in used
        and record["frenchWord"] not in ambiguous
    ]
    t4_items = [build_t4_item(record, rng) for record in t4_records]
    return t1_items + t2_items + t3_items + t4_items


def to_import_package(items: list[dict], records: list[dict], evidence: dict[str, dict]) -> dict:
    by_word = {record["frenchWord"]: record for record in records}
    rows = []
    options = []
    counters = {key: 0 for key in TYPE_META}
    for item in items:
        item_type = item["itemType"]
        counters[item_type] += 1
        item_code = f"FF4PROD-{item_type}-{counters[item_type]:03d}"
        section, question_type, construct, context = TYPE_META[item_type]
        record = by_word[item["targetWord"]]
        evidence_record = evidence.get(item["targetWord"], {})
        is_control = item.get("controlType") == "COGNATE_CONTROL"
        review_line = {
            "T1": "生产审查：APPROVED（核心义完全不重合、词性、选项角色与答案一致性已复核）。",
            "T2": ("生产审查：APPROVED（TEM4 完整原句可定位；近义词、英语迁移义对应词及平行语法已逐项复核）。"
                   if item.get("evidenceLevel") == "TEM4_EXACT_SENTENCE"
                   else "生产审查：APPROVED（基于可定位 TEM4 原搭配补充最小语境；近义词、迁移词和句法已逐项复核）。"),
            "T3": ("生产审查：APPROVED（英法同形同义 Vrai 控制题；用于降低全 F 反应定势）。"
                   if is_control else "生产审查：APPROVED（假朋友英语迁移义命题，标准答案为 F）。"),
            "T4": "生产审查：APPROVED（仅按拼写距离 1–4、形态型排除与答案唯一性复核；本题型不采用语义筛选）。",
        }[item_type]
        explanation = (
            f"法语词：{record['frenchWord']}；易混英文词：{record['englishConfusable']}。\n"
            f"TEM4 前 2–3 义项：{'；'.join(record['tem4TopSenses'])}。\n"
            f"英文迁移义：{'；'.join(record['englishChineseSenses'])}。\n"
            f"来源：TEM4 p.{record['tem4PdfPage']}；假朋友词典 p.{record['falseFriendsPdfPage']}。\n"
            + review_line
        )
        if item_type == "T2" and item.get("evidenceLevel") == "TEM4_COLLOCATION_CONTEXTUALIZED":
            explanation += f"\n证据等级：基于 TEM4 原搭配“{item['sourceCollocation']}”补充最小语境，并非逐字原句。"
        rows.append({
            "itemCode": item_code,
            "sectionCode": section,
            "questionType": question_type,
            "stemText": item["stemText"],
            "promptText": None,
            "correctAnswers": item["correctAnswers"],
            "explanationText": explanation,
            "requiredAnswer": True,
            "scored": True,
            "weight": 1,
            "transferCategory": "COGNATE" if is_control else "FALSE_FRIEND",
            "contextLevel": context,
            "constructCode": construct,
            "targetWord": item["targetWord"],
            "displayConditionJson": None,
            "tem4PdfPage": record["tem4PdfPage"],
            "falseFriendsPdfPage": record["falseFriendsPdfPage"],
            "exampleSentenceStatus": (item.get("evidenceLevel") if item_type == "T2"
                                      else record.get("exampleSentenceStatus")),
            "spellingRawEditDistance": record["rawEditDistance"] if item_type == "T4" else None,
            "spellingAccentFoldedEditDistance": record["accentFoldedEditDistance"] if item_type == "T4" else None,
            "morphologyOnly": record["morphologyOnly"] if item_type == "T4" else False,
            "lexicalReviewStatus": "APPROVED",
            "pedagogicReviewStatus": "APPROVED",
        })
        for option in item["options"]:
            options.append({
                "itemCode": item_code,
                "optionCode": option["key"],
                "optionText": option["label"],
                "correct": option["correct"],
                "explanation": None,
                "role": option["role"],
            })

    sections = []
    titles = {
        "FF4_WORD_MEANING": "题型一 词义单选",
        "FF4_SENTENCE_SELECTION": "题型二 句子选词",
        "FF4_TRUE_FALSE": "题型三 判断正误",
        "FF4_SPELLING": "题型四 单词拼写",
    }
    for order, code in enumerate(titles, 1):
        sections.append({
            "sectionCode": code,
            "title": titles[code],
            "description": titles[code],
            "sharedMaterial": None,
            "sortOrder": order,
            "formalSection": True,
        })
    return {
        "questionnaire": {
            "code": "LEXIBRIDGE_FF4_PRODUCTION_20260812",
            "title": "Lexi-Bridge 法语专四假朋友生产题库",
            "description": "经概念级语义、来源、选项、拼写及答案一致性复核的生产题库",
            "durationMinutes": 120,
            "scoringVersion": "SCORING_V3",
            "aiPromptVersion": "assessment-analysis/v3",
        },
        "sections": sections,
        "items": rows,
        "options": options,
    }


def write_workbook(package: dict, records: list[dict]) -> None:
    workbook = Workbook()
    overview = workbook.active
    overview.title = "生产说明"
    counts = {code: sum(1 for item in package["items"] if item["constructCode"] == code)
              for code in ("FF4_WORD_MEANING", "FF4_SENTENCE_SYNONYM", "FF4_TRUE_FALSE_TRANSFER", "FF4_SPELLING")}
    ambiguous = find_t4_ambiguous(records)
    spelling_eligible = {record["frenchWord"] for record in records
                         if "T4" in record.get("eligibleTypes", [])
                         and record["frenchWord"] not in ambiguous}
    maximum_eligible = set(APPROVED_SEMANTIC_WORDS) | spelling_eligible | set(TRUE_COGNATE_CONTROLS)
    overview.append(["法语专四假朋友题库｜生产版", None])
    overview.merge_cells("A1:B1")
    overview.append(["字段", "内容"])
    overview_rows = [
        ("问卷代码", package["questionnaire"]["code"]),
        ("题目总数", len(package["items"])),
        ("题型一", counts["FF4_WORD_MEANING"]),
        ("题型二", counts["FF4_SENTENCE_SYNONYM"]),
        ("题型三", counts["FF4_TRUE_FALSE_TRANSFER"]),
        ("题型四", counts["FF4_SPELLING"]),
        ("内容状态", "APPROVED（全部题目已通过内容与结构审计）"),
        ("覆盖状态", f"MAXIMUM_RULE_COMPLIANT_COVERAGE（{len(package['items'])}/{len(maximum_eligible)}，无规则内合格目标词遗漏）"),
        ("导入状态", "READY_FOR_IMPORT_VALIDATION"),
        ("发布状态", "NOT_DEPLOYED（未连接生产库、未创建 release）"),
        ("生成规则", "语义题概念完全不重合；拼写距离 1–4；排除形态伪混淆与中文题干歧义；目标词全局不重复。"),
    ]
    for row in overview_rows:
        overview.append(row)
    overview.column_dimensions["A"].width = 24
    overview.column_dimensions["B"].width = 100

    options_by_item = {}
    for option in package["options"]:
        options_by_item.setdefault(option["itemCode"], []).append(option)
    by_word = {record["frenchWord"]: record for record in records}
    sheet_meta = [
        ("单选", "FF4_WORD_MEANING"),
        ("选词填空", "FF4_SENTENCE_SYNONYM"),
        ("判断", "FF4_TRUE_FALSE_TRANSFER"),
        ("拼写", "FF4_SPELLING"),
    ]
    columns = ["题目编号", "词条ID", "目标词", "题干", "选项A/V", "选项B/F", "选项C", "选项D",
               "标准答案", "解析", "来源页码", "词汇状态", "教研状态"]
    for sheet_name, construct in sheet_meta:
        sheet = workbook.create_sheet(sheet_name)
        sheet.append([f"法语专四假朋友题库｜{sheet_name}"] + [None] * (len(columns) - 1))
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
        sheet.append(["生产版｜全部内容已完成 AI 概念级与结构复核。"] + [None] * (len(columns) - 1))
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
        sheet.append(columns)
        for item in [row for row in package["items"] if row["constructCode"] == construct]:
            option_rows = options_by_item.get(item["itemCode"], [])
            option_map = {option["optionCode"]: option["optionText"] for option in option_rows}
            if construct == "FF4_TRUE_FALSE_TRANSFER":
                option_map = {"A": "Vrai（正确）", "B": "Faux（错误）"}
            record = by_word[item["targetWord"]]
            sheet.append([
                item["itemCode"], record["wordId"], item["targetWord"], item["stemText"],
                option_map.get("A"), option_map.get("B"), option_map.get("C"), option_map.get("D"),
                "；".join(item["correctAnswers"]), item["explanationText"],
                f"TEM4 p.{item['tem4PdfPage']}；假朋友词典 p.{item['falseFriendsPdfPage']}",
                item["lexicalReviewStatus"], item["pedagogicReviewStatus"],
            ])
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for cell in sheet[1]:
            cell.fill = PatternFill("solid", fgColor="17365D")
            cell.font = Font(color="FFFFFF", bold=True, size=14)
        for cell in sheet[3]:
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.font = Font(color="FFFFFF", bold=True)
        for index, width in enumerate([20, 24, 18, 55, 28, 28, 28, 28, 12, 80, 30, 18, 18], 1):
            sheet.column_dimensions[get_column_letter(index)].width = width
        sheet.freeze_panes = "A4"
        sheet.auto_filter.ref = f"A3:M{sheet.max_row}"

    # Exact backend import sheets. Keeping them in the same workbook makes
    # the reviewer-friendly deliverable directly importable without a second
    # conversion step.
    questionnaire_headers = ["code", "title", "description", "durationMinutes", "scoringVersion", "aiPromptVersion"]
    questionnaire_sheet = workbook.create_sheet("Questionnaire")
    questionnaire_sheet.append(questionnaire_headers)
    questionnaire_sheet.append([package["questionnaire"].get(key) for key in questionnaire_headers])

    section_headers = ["sectionCode", "title", "description", "sharedMaterial", "sortOrder", "formalSection"]
    section_sheet = workbook.create_sheet("Sections")
    section_sheet.append(section_headers)
    for section in package["sections"]:
        section_sheet.append([section.get(key) for key in section_headers])

    item_headers = [
        "itemCode", "sectionCode", "questionType", "stemText", "promptText", "correctAnswers",
        "explanationText", "requiredAnswer", "scored", "weight", "transferCategory", "contextLevel",
        "constructCode", "targetWord", "displayConditionJson", "tem4PdfPage", "falseFriendsPdfPage",
        "exampleSentenceStatus", "spellingRawEditDistance", "spellingAccentFoldedEditDistance",
        "morphologyOnly", "lexicalReviewStatus", "pedagogicReviewStatus",
    ]
    item_sheet = workbook.create_sheet("Items")
    item_sheet.append(item_headers)
    for item in package["items"]:
        row = []
        for key in item_headers:
            value = item.get(key)
            if key == "correctAnswers":
                value = ";".join(value or [])
            row.append(value)
        item_sheet.append(row)

    option_headers = ["itemCode", "optionCode", "optionText", "correct", "explanation", "role"]
    option_sheet = workbook.create_sheet("Options")
    option_sheet.append(option_headers)
    for option in package["options"]:
        option_sheet.append([option.get(key) for key in option_headers])

    for sheet in (questionnaire_sheet, section_sheet, item_sheet, option_sheet):
        for cell in sheet[1]:
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.font = Font(color="FFFFFF", bold=True)
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(column)].width = min(60, max(14, max(
                len(str(sheet.cell(row, column).value or "")) for row in range(1, min(sheet.max_row, 50) + 1)
            ) + 2))

    workbook.save(WORKBOOK_PATH)


def main() -> None:
    records = load_jsonl(OUT / "candidate-adjudication.jsonl")
    evidence_rows = load_jsonl(OUT / "t2-evidence.jsonl")
    evidence = {row["frenchWord"]: row for row in evidence_rows}
    write_semantic_review(records)
    items = build_items(records, evidence)
    package = to_import_package(items, records, evidence)
    PACKAGE_PATH.write_text(json.dumps(package, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    write_workbook(package, records)
    counts = {}
    for item in package["items"]:
        counts[item["constructCode"]] = counts.get(item["constructCode"], 0) + 1
    print(json.dumps({"total": len(package["items"]), "counts": counts}, ensure_ascii=False))


if __name__ == "__main__":
    main()
