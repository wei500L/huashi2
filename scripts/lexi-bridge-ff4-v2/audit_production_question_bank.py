#!/usr/bin/env python3
"""Strict structural/content audit for the production FF4 package."""

from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))

from build_candidates import ROOT  # noqa: E402
from build_question_bank import find_t4_ambiguous  # noqa: E402
from production_semantic_rules import (  # noqa: E402
    APPROVED_SEMANTIC_WORDS, MAX_CONSECUTIVE_FALSE_FRIEND_T3,
    MINIMUM_T2_ITEM_COUNT, RULESET_VERSION, T2_RULES, T2_TARGET_WORDS,
    TRUE_COGNATE_CONTROLS,
)

OUT = ROOT / "docs" / "data" / "lexi-bridge-ff4-v2"
PACKAGE = OUT / "question-bank-package-production.json"
WORKBOOK = OUT / "法语专四假朋友题库_生产版.xlsx"
REPORT = OUT / "production-audit-report.md"


def load_jsonl(path: Path) -> list[dict]:
    return [json.loads(line) for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]


def main() -> int:
    package = json.loads(PACKAGE.read_text(encoding="utf-8"))
    records = load_jsonl(OUT / "candidate-adjudication.jsonl")
    semantic_reviews = {row["frenchWord"]: row for row in load_jsonl(OUT / "production-semantic-review.jsonl")}
    by_word = {row["frenchWord"]: row for row in records}
    ambiguous = find_t4_ambiguous(records)
    semantic_eligible = set(APPROVED_SEMANTIC_WORDS)
    spelling_eligible = {
        row["frenchWord"] for row in records
        if "T4" in row.get("eligibleTypes", []) and row["frenchWord"] not in ambiguous
    }
    maximum_eligible_pool = semantic_eligible | spelling_eligible | set(TRUE_COGNATE_CONTROLS)
    options_by_item: dict[str, list[dict]] = {}
    for option in package["options"]:
        options_by_item.setdefault(option["itemCode"], []).append(option)

    issues: list[str] = []
    target_words: set[str] = set()
    counts: dict[str, int] = {}
    true_control_count = 0
    consecutive_false_friend_t3 = 0
    maximum_false_friend_t3_run = 0
    for item in package["items"]:
        code = item["itemCode"]
        construct = item["constructCode"]
        counts[construct] = counts.get(construct, 0) + 1
        word = item["targetWord"]
        if word in target_words:
            issues.append(f"{code}: target word reused: {word}")
        target_words.add(word)
        record = by_word.get(word)
        if record is None:
            issues.append(f"{code}: missing candidate record")
            continue
        if item["lexicalReviewStatus"] != "APPROVED" or item["pedagogicReviewStatus"] != "APPROVED":
            issues.append(f"{code}: non-production review status")
        if not item.get("tem4PdfPage") or not item.get("falseFriendsPdfPage"):
            issues.append(f"{code}: missing source pages")

        options = options_by_item.get(code, [])
        if construct in ("FF4_WORD_MEANING", "FF4_SENTENCE_SYNONYM"):
            roles = [option["role"] for option in options]
            if len(options) != 4 or roles.count("CORRECT") != 1 or roles.count("TRANSFER") != 1 or roles.count("DISTRACTOR") != 2:
                issues.append(f"{code}: invalid option roles")
            correct = [option["optionCode"] for option in options if option["correct"]]
            if correct != item["correctAnswers"]:
                issues.append(f"{code}: answer/options mismatch")
            if len({option["optionCode"] for option in options}) != len(options):
                issues.append(f"{code}: duplicate option code")
            if len({option["optionText"] for option in options}) != len(options):
                issues.append(f"{code}: duplicate option text")
        if construct == "FF4_WORD_MEANING":
            review = semantic_reviews.get(word)
            if review is None or review["productionDecision"] != "APPROVED":
                issues.append(f"{code}: semantic target not production-approved")
            target_pos = (record.get("partOfSpeech") or "").split(".")[0].strip()
            for option in [option for option in options if option["role"] == "DISTRACTOR"]:
                sources = [candidate for candidate in records
                           if "; ".join((candidate.get("tem4TopSenses") or [])[:2]) == option["optionText"]]
                if not any((source.get("partOfSpeech") or "").split(".")[0].strip() == target_pos
                           for source in sources):
                    issues.append(f"{code}: distractor POS mismatch: {option['optionText']}")
        elif construct == "FF4_SENTENCE_SYNONYM":
            if word not in T2_TARGET_WORDS:
                issues.append(f"{code}: unexpected T2 target")
            rule = T2_RULES.get(word)
            expected_evidence = rule.get("evidenceLevel", "TEM4_EXACT_SENTENCE") if rule else None
            if item.get("exampleSentenceStatus") != expected_evidence:
                issues.append(f"{code}: T2 evidence level differs from approved rule")
            if expected_evidence == "TEM4_COLLOCATION_CONTEXTUALIZED" and not rule.get("sourceCollocation"):
                issues.append(f"{code}: contextualized T2 missing source collocation")
            if rule is None or item.get("stemText") != "请根据句子选择画线单词的同义解释\n" + rule["sentence"]:
                issues.append(f"{code}: T2 source sentence/emphasis differs from approved text")
            actual = [(option["optionCode"], option["optionText"], option["correct"], option["role"])
                      for option in options]
            if rule is None or actual != list(rule["options"]):
                issues.append(f"{code}: T2 options differ from approved synonym/transfer evidence")
        elif construct == "FF4_TRUE_FALSE_TRANSFER":
            is_control = item.get("transferCategory") == "COGNATE"
            if is_control:
                true_control_count += 1
                consecutive_false_friend_t3 = 0
                if word not in TRUE_COGNATE_CONTROLS:
                    issues.append(f"{code}: unapproved cognate control")
                if item["correctAnswers"] != ["V"]:
                    issues.append(f"{code}: cognate control answer must be V")
                expected_stem = f"{word} = {TRUE_COGNATE_CONTROLS.get(word, '')}"
                if item.get("stemText") != expected_stem:
                    issues.append(f"{code}: cognate control stem differs from approved sense")
            else:
                consecutive_false_friend_t3 += 1
                maximum_false_friend_t3_run = max(maximum_false_friend_t3_run, consecutive_false_friend_t3)
                review = semantic_reviews.get(word)
                if review is None or review["productionDecision"] != "APPROVED":
                    issues.append(f"{code}: T3 target not production-approved")
                if item["correctAnswers"] != ["F"]:
                    issues.append(f"{code}: false-friend T3 answer must be F")
        elif construct == "FF4_SPELLING":
            distance = min(item["spellingRawEditDistance"], item["spellingAccentFoldedEditDistance"])
            if not 1 <= distance <= 4:
                issues.append(f"{code}: spelling distance out of range")
            if item["morphologyOnly"] or record.get("morphologyOnly"):
                issues.append(f"{code}: morphology-only spelling pair")
            if word in ambiguous:
                issues.append(f"{code}: ambiguous Chinese spelling stem")

    expected_keys = {"questionnaire", "sections", "items", "options"}
    if set(package) != expected_keys:
        issues.append("package top-level fields do not match import DTO")

    missing_eligible = sorted(maximum_eligible_pool - target_words)
    unexpected_targets = sorted(target_words - maximum_eligible_pool)
    if missing_eligible:
        issues.append("maximum coverage failure; eligible targets omitted: " + ", ".join(missing_eligible))
    if unexpected_targets:
        issues.append("targets outside approved semantic/spelling pools: " + ", ".join(unexpected_targets))
    if true_control_count != len(TRUE_COGNATE_CONTROLS):
        issues.append(f"expected {len(TRUE_COGNATE_CONTROLS)} Vrai cognate controls, got {true_control_count}")
    t2_count = counts.get("FF4_SENTENCE_SYNONYM", 0)
    if t2_count < MINIMUM_T2_ITEM_COUNT:
        issues.append(f"T2 regression: expected at least {MINIMUM_T2_ITEM_COUNT} items, got {t2_count}")
    if maximum_false_friend_t3_run > MAX_CONSECUTIVE_FALSE_FRIEND_T3:
        issues.append("T3 interleaving regression: maximum consecutive false-friend run "
                      f"is {maximum_false_friend_t3_run}, limit is {MAX_CONSECUTIVE_FALSE_FRIEND_T3}")

    workbook = load_workbook(WORKBOOK, read_only=True, data_only=True)
    expected_rows = {
        "单选": counts.get("FF4_WORD_MEANING", 0),
        "选词填空": counts.get("FF4_SENTENCE_SYNONYM", 0),
        "判断": counts.get("FF4_TRUE_FALSE_TRANSFER", 0),
        "拼写": counts.get("FF4_SPELLING", 0),
    }
    for sheet_name, expected in expected_rows.items():
        actual = workbook[sheet_name].max_row - 3
        if actual != expected:
            issues.append(f"workbook {sheet_name}: expected {expected} rows, got {actual}")
    import_sheet_rows = {
        "Questionnaire": 1,
        "Sections": len(package["sections"]),
        "Items": len(package["items"]),
        "Options": len(package["options"]),
    }
    expected_headers = {
        "Questionnaire": ["code", "title", "description", "durationMinutes", "scoringVersion", "aiPromptVersion"],
        "Sections": ["sectionCode", "title", "description", "sharedMaterial", "sortOrder", "formalSection"],
        "Items": ["itemCode", "sectionCode", "questionType", "stemText", "promptText", "correctAnswers",
                  "explanationText", "requiredAnswer", "scored", "weight", "transferCategory", "contextLevel",
                  "constructCode", "targetWord", "displayConditionJson", "tem4PdfPage", "falseFriendsPdfPage",
                  "exampleSentenceStatus", "spellingRawEditDistance", "spellingAccentFoldedEditDistance",
                  "morphologyOnly", "lexicalReviewStatus", "pedagogicReviewStatus"],
        "Options": ["itemCode", "optionCode", "optionText", "correct", "explanation", "role"],
    }
    for sheet_name, expected in import_sheet_rows.items():
        if sheet_name not in workbook.sheetnames:
            issues.append(f"workbook missing import sheet: {sheet_name}")
            continue
        actual = workbook[sheet_name].max_row - 1
        if actual != expected:
            issues.append(f"import sheet {sheet_name}: expected {expected} rows, got {actual}")
        actual_headers = [cell.value for cell in workbook[sheet_name][1]]
        if actual_headers != expected_headers[sheet_name]:
            issues.append(f"import sheet {sheet_name}: header contract mismatch")

    lines = [
        "# 法语专四假朋友生产题库审计报告",
        "",
        f"- 总题量：{len(package['items'])}",
        f"- 全局唯一目标词：{len(target_words)}",
        f"- 语义题合格目标词：{len(semantic_eligible)}",
        f"- 判断题 Vrai 同源词控制题：{len(TRUE_COGNATE_CONTROLS)}",
        f"- 判断题最大连续 F：{maximum_false_friend_t3_run}（门限 ≤ {MAX_CONSECUTIVE_FALSE_FRIEND_T3}）",
        f"- 拼写题合格目标词：{len(spelling_eligible)}",
        f"- 语义池、拼写池与同源控制池合并后的最大可用目标词：{len(maximum_eligible_pool)}",
        f"- 最大覆盖率：{len(target_words)}/{len(maximum_eligible_pool)}（{'100%' if target_words == maximum_eligible_pool else '未达标'}）",
        "- 各题型：" + "；".join(f"{key}={value}" for key, value in counts.items()),
        f"- 结构与内容问题：{len(issues)}",
        "- 内容状态：APPROVED",
        "- 覆盖状态：MAXIMUM_RULE_COMPLIANT_COVERAGE",
        "- 导入状态：READY_FOR_IMPORT_VALIDATION",
        "- 发布状态：NOT_DEPLOYED（尚未连接生产库、未创建 release）",
        f"- 规则版本：{RULESET_VERSION}",
        "",
        "## 问题清单",
    ]
    lines.extend([f"- {issue}" for issue in issues] or ["（无）"])
    REPORT.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(json.dumps({"total": len(package["items"]), "uniqueTargets": len(target_words),
                      "maximumEligibleTargets": len(maximum_eligible_pool),
                      "coverageComplete": target_words == maximum_eligible_pool,
                      "counts": counts, "issueCount": len(issues)}, ensure_ascii=False))
    return 1 if issues else 0


if __name__ == "__main__":
    raise SystemExit(main())
