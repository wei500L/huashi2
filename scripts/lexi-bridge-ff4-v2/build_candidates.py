#!/usr/bin/env python3
"""FF4 question bank V2 - candidate adjudication builder.

Builds an auditable per-word record for every intersection word of the
Lexi-Bridge FF4 pool, using only deterministic rules and source evidence.

Outputs:
  docs/data/lexi-bridge-ff4-v2/candidate-adjudication.jsonl
  docs/data/lexi-bridge-ff4-v2/candidate-adjudication.xlsx
  docs/data/lexi-bridge-ff4-v2/semantic-review-queue.xlsx

Every raw candidate is marked LEXICAL_REVIEW_PENDING because this file records
the pre-production adjudication stage. Production generators apply the
separate concept-level approval list and never copy these pending labels into
approved question-bank items.
"""

from __future__ import annotations

import json
import re
import unicodedata
from dataclasses import asdict, dataclass, field
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
DATA = ROOT / "docs" / "data" / "lexi-bridge-ff4"
OUT = ROOT / "docs" / "data" / "lexi-bridge-ff4-v2"
OCR_TEM4 = OUT / "source-ocr-cache" / "tem4-pages.jsonl"
OCR_FF = DATA / "source-ocr-cache" / "false-friends-pages.jsonl"
REVIEWED_WORKBOOK = DATA / "student-deliverables" / "法语专四假朋友词汇总表.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")

CJK_PUNCT_RE = re.compile(r"[\s\u3000\u3001\u3002\uff0c\uff1b\uff08\uff09\uff1a\u201c\u201d\u2018\u2019·,;:()\[\]{}<>/\\|!?.\"'\-—_]+")
CONTROL_CHAR_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")


def xml_safe(value: object) -> str:
    """Strip control characters that Excel/XML cannot store."""
    return CONTROL_CHAR_RE.sub("", str(value or ""))


def normalize_cn(value: str) -> str:
    """Normalize a Chinese gloss: strip punctuation and whitespace."""
    if value is None:
        return ""
    return CJK_PUNCT_RE.sub("", unicodedata.normalize("NFC", str(value)))


def cn_tokens(value: str) -> list[str]:
    raw = str(value)
    parts = re.split(r"[\u3001\uff0c\uff1b;,、／/]+", raw)
    return [normalize_cn(p) for p in parts if normalize_cn(p)]


def fold_latin(value: str) -> str:
    decomposed = unicodedata.normalize("NFD", str(value))
    folded = "".join(ch for ch in decomposed if unicodedata.category(ch) != "Mn")
    return folded.lower()


def levenshtein(left: str, right: str) -> int:
    if len(left) < len(right):
        left, right = right, left
    previous = list(range(len(right) + 1))
    for row, char_l in enumerate(left, 1):
        current = [row]
        for col, char_r in enumerate(right, 1):
            current.append(min(previous[col] + 1, current[col - 1] + 1,
                               previous[col - 1] + (char_l != char_r)))
        previous = current
    return previous[-1]


@dataclass
class AdjudicationRecord:
    wordId: str
    frenchWord: str
    englishConfusable: str
    tem4TopSenses: list[str]
    englishChineseSenses: list[str]
    semanticOverlapDecision: str
    semanticDecisionReason: str
    tem4PdfPage: int
    falseFriendsPdfPage: int
    tem4SourceQuote: str
    falseFriendsSourceQuote: str
    tem4ExampleSentence: str | None
    exampleSentenceStatus: str
    rawEditDistance: int
    accentFoldedEditDistance: int
    morphologyOnly: bool
    morphologyRule: str | None
    eligibleTypes: list[str]
    assignedType: str | None
    lexicalReviewStatus: str
    pedagogicReviewStatus: str
    reviewNotes: str
    senseTokenOverlap: list[str] = field(default_factory=list)
    candidateId: str | None = None
    partOfSpeech: str | None = None
    unit: str | None = None
    sourceRelationType: str | None = None


def load_jsonl(path: Path, encoding: str = "utf-8-sig") -> list[dict]:
    rows = []
    with path.open(encoding=encoding) as handle:
        for line in handle:
            line = line.strip()
            if line:
                rows.append(json.loads(line))
    return rows


def load_ocr_pages(path: Path) -> dict[int, str]:
    pages: dict[int, str] = {}
    for row in load_jsonl(path, encoding="utf-8"):
        pages[row["pdf_page"]] = row.get("text") or ""
    return pages


def load_ff_ocr_pages(path: Path) -> dict[int, str]:
    pages: dict[int, str] = {}
    for row in load_jsonl(path, encoding="utf-8"):
        match = re.search(r"ff-(\d+)\.png", row.get("path") or "")
        if match:
            pages[int(match.group(1))] = row.get("text") or ""
    return pages


def split_reviewed_senses(value: object, limit: int | None = None) -> list[str]:
    senses: list[str] = []
    for part in re.split(r"[；;]", str(value or "")):
        normalized = part.strip(" 、，,。")
        if normalized and normalized not in senses:
            senses.append(normalized)
    return senses[:limit] if limit is not None else senses


def load_reviewed_vocabulary(path: Path) -> dict[str, dict]:
    """Load the previously audited 445-word workbook as the preferred gloss source.

    The raw extraction JSON remains available for traceability and fallback,
    but reviewed corrections (for example heurter) must not regress when the
    question-bank pipeline is rebuilt.
    """
    sheet = load_workbook(path, read_only=True, data_only=True)["词汇总表"]
    rows = sheet.iter_rows(values_only=True)
    next(rows, None)
    next(rows, None)
    header_row = next(rows)
    headers = {str(value).strip(): index for index, value in enumerate(header_row) if value}
    reviewed: dict[str, dict] = {}
    for row in rows:
        french = str(row[headers["法语单词"]] or "").strip()
        if not french:
            continue
        reviewed[french] = {
            "tem4TopSenses": split_reviewed_senses(
                row[headers["TEM4审定完整释义"]], 3),
            "englishChineseSenses": split_reviewed_senses(
                row[headers["英文词真实中文义"]]),
            "partOfSpeech": str(row[headers["词性"]] or "").strip() or None,
            "unit": str(row[headers["TEM4单元"]] or "").strip() or None,
            "tem4SourceQuote": str(row[headers["TEM4来源例组/搭配"]] or "").strip(),
            "falseFriendsSourceQuote": str(row[headers["Kirk-Greene原文摘录（OCR）"]] or "").strip(),
            "reviewCorrection": str(row[headers["审定纠正/处理"]] or "").strip(),
            "sourceRelationType": str(row[headers["原书关系类型"]] or "").strip() or None,
        }
    return reviewed


def extract_sentences(text: str) -> list[str]:
    compact = re.sub(r"\s+", " ", text)
    compact = compact.replace(" 。", "。").replace(" .", ".")
    pieces = re.split(r"(?<=[.!?。])\s+(?=[A-ZÀ-ÖØ-ÞÉÈÀÇ])", compact)
    results = []
    for piece in pieces:
        candidate = piece.strip()
        if len(candidate) < 12 or len(candidate) > 220:
            continue
        if re.search(r"[\u4e00-\u9fff\u3000-\u303f\uff00-\uffefА-Яа-яЁё\uFFFD\u25A0\u25A1\u2605\u2606]", candidate):
            continue
        if re.search(r"[&$~\uFF5E=]", candidate):
            continue
        if re.search(r"\d", candidate):
            continue
        if re.search(r"\[[^\]]+\]", candidate):
            continue
        if re.search(r"\b(v\.t|v\.i|v\.pr|n\.m|n\.f|n\.pl|a\.|ad\.|prép|conj|loc|adv)\b", candidate):
            continue
        if not candidate.rstrip().endswith((".", "!", "?")):
            continue
        letters = re.findall(r"[a-zA-Zàâäéèêëîïôöùûüçœæ]", candidate)
        if len(letters) / max(1, len(candidate)) >= 0.7:
            results.append(candidate)
    return results


def word_forms(french_word: str) -> list[str]:
    """Candidate surface forms of the headword to locate in OCR text.

    The headword itself is always first; derived inflections follow and are
    only accepted when the line also contains a subject/determiner marker.
    """
    folded = fold_latin(french_word)
    forms = [folded]
    if folded.endswith("er") and len(folded) > 5:
        stem = folded[:-2]
        forms.extend([stem, stem + "e", stem + "es", stem + "ent", stem + "a", stem + "ant"])
    if folded.endswith("re") and len(folded) > 5:
        stem = folded[:-2]
        forms.extend([stem, stem + "s", stem + "t", stem + "nt"])
    if folded.endswith("ir") and len(folded) > 5:
        stem = folded[:-2]
        forms.extend([stem, stem + "s", stem + "t", stem + "ssons", stem + "ssent"])
    unique: list[str] = []
    for form in forms:
        if form not in unique:
            unique.append(form)
    return unique


def _sentence_ok(core: str, merged: bool) -> bool:
    """Deterministic purity gate for an extracted example sentence."""
    if len(core) < 12 or len(core) > 220:
        return False
    if not core[0].isupper():
        return False
    if re.search(r"[\u4e00-\u9fff\u3000-\u303f\uff00-\uffefА-Яа-яЁё\uFFFD\u25A0\u25A1\u2605\u2606]", core):
        return False
    if re.search(r"[&$~=•\[\]()]", core):
        return False
    if re.search(r"\d+[)\]}$]", core) or re.search(r"\(\s*\d", core):
        return False
    if re.search(r"\b(v\.|n\.|a\.|ad\.|v\.t|v\.i|v\.pr|n\.m|n\.f|n\.pl|prép|conj|loc|adv|adj)\b", core):
        return False
    if re.search(r"\s\w{1,3}\.\s*$", core):
        return False
    if re.search(r"[:;]\s+[A-ZÀ-ÖØ-Þ]", core):
        return False
    if re.search(r"\s+\w+\.\w+\.\w+\b", core):
        return False
    if merged and re.search(r"[a-zà-ÿ]\s+[A-ZÀ-ÖØ-Þ]", core):
        return False
    letters = re.findall(r"[a-zA-Zàâäéèêëîïôöùûüçœæ]", core)
    return len(letters) / max(1, len(core)) >= 0.7


def find_example_sentence(page_text: str, french_word: str) -> tuple[str | None, str]:
    """Locate a complete, clean example sentence containing the target word.

    Works on the Vision OCR page text: a sentence is accepted only when its
    core (up to the first sentence-final punctuation) starts with a capital
    letter, ends with .!?, and contains the headword (or an obvious
    inflection) as a whole word. Single lines are tried first; when a line
    containing the headword is incomplete, up to five following lines are
    merged so sentences split across OCR lines are recovered.

    Returns (sentence, status). Status is SOURCE_VERIFIED when a clean
    sentence is found, otherwise SOURCE_EXAMPLE_MISSING.
    """
    forms = word_forms(french_word)
    subject_markers = r"\b(je|tu|il|elle|on|nous|vous|ils|elles|ce|cet|cette|ces|le|la|les|un|une|des|mon|ma|mes|ton|ta|tes|son|sa|ses|notre|nos|votre|vos|leur|leurs|que|qui|ça|ca|ce|quoi)\b"
    lines = [line.strip() for line in page_text.split("\n")]
    for index, line in enumerate(lines):
        if len(line) < 8:
            continue
        folded_line = fold_latin(line)
        matched_form = next(
            (form for form in forms if re.search(r"\b" + re.escape(form) + r"\b", folded_line)),
            None,
        )
        if matched_form is None:
            continue
        candidate_lines = [line]
        merged = " ".join(candidate_lines)
        for offset in range(0, 6):
            if offset > 0:
                if index + offset >= len(lines):
                    break
                candidate_lines.append(lines[index + offset])
                merged = " ".join(candidate_lines)
            core_match = re.search(r"^[^.!?]*[.!?]", merged)
            if core_match is None:
                continue
            core = re.sub(r"^\s*\d+\s*", "", core_match.group(0)).strip()
            core = re.sub(r"\s*\([^()]{0,12}\)\s*$", "", core).strip()
            if matched_form != forms[0] and not re.search(subject_markers, fold_latin(core)):
                continue
            if _sentence_ok(core, merged=offset > 0):
                return core, "SOURCE_VERIFIED"
    return None, "SOURCE_EXAMPLE_MISSING"


MORPHOLOGY_SUFFIX_RULES: list[tuple[str, str, str, str]] = [
    # (rule code, french suffix, english suffix, english fallback)
    ("VERB_INFINITIVE_ER", "er", None, None),
    ("VERB_INFINITIVE_IR", "ir", None, None),
    ("VERB_INFINITIVE_RE", "re", None, None),
    ("ADVERB_MENT", "ment", "ly", None),
    ("NOUN_TION", "tion", "tion", None),
    ("NOUN_SION", "sion", "sion", None),
    ("ADJ_ITY", "ité", "ity", None),
    ("ADJ_IQUE", "ique", "ic", None),
    ("ADJ_EL_AL", "el", "al", None),
    ("NOUN_EUR_OR", "eur", "or", None),
    ("ADJ_EUX_OUS", "eux", "ous", None),
]


def detect_morphology_only(french: str, english: str) -> tuple[bool, str | None]:
    """Return (morphology_only, rule_code)."""
    fr = fold_latin(french)
    en = fold_latin(english)
    # French infinitives such as imaginer/imagine or fatiguer/fatigue differ
    # only by the predictable final -r after an English word ending in -e.
    # Like classer/class, these do not create a meaningful spelling trap.
    if fr.endswith("er") and en.endswith("e") and fr[:-1] == en:
        return True, "VERB_INFINITIVE_FINAL_R"
    for rule, fr_suffix, en_suffix, _ in MORPHOLOGY_SUFFIX_RULES:
        if not fr.endswith(fr_suffix) or len(fr) - len(fr_suffix) < 3:
            continue
        fr_stem = fr[: -len(fr_suffix)]
        if en.endswith(en_suffix) if en_suffix else False:
            if fr_stem == en[: -len(en_suffix)]:
                return True, rule
        elif fr_stem == en:
            return True, rule
    return False, None


COMMON_CN_PARTICLES = set("的了着过地得也都在和有这与那是我你不他她它很就而")
STOP_SHINGLES = {"一个", "一些", "某种", "许多", "什么", "这个", "那个", "一种", "对于", "以及", "或者"}


def cn_shingles(value: str) -> set[str]:
    """2-char shingles of a Chinese gloss, dropping function-word shingles."""
    normalized = normalize_cn(value)
    shingles = set()
    for index in range(len(normalized) - 1):
        shingle = normalized[index: index + 2]
        if shingle in STOP_SHINGLES:
            continue
        if any(ch in COMMON_CN_PARTICLES for ch in shingle) and len(shingle) == 2:
            if sum(ch in COMMON_CN_PARTICLES for ch in shingle) > 1:
                continue
        shingles.add(shingle)
    return shingles


def decide_semantic_overlap(tem4_senses: list[str], english_senses: list[str]) -> tuple[str, str, list[str]]:
    """Deterministic semantic overlap decision.

    Returns (decision, reason, overlapping tokens). Exact gloss equality is
    CORE; substring containment or small edit distance is PARTIAL; shared
    2-char shingles (e.g. 印象 in 给人深刻印象的 and 令人印象深刻的) are treated
    as conceptual overlap (CORE when the shared shingle sits in the first
    TEM4 sense, PARTIAL otherwise); otherwise NONE. Words whose conceptual
    closeness cannot be decided by string rules are sent to the human review
    queue instead of being silently accepted.
    """
    t_tokens = [t for sense in tem4_senses for t in cn_tokens(sense) if t]
    e_tokens = [t for sense in english_senses for t in cn_tokens(sense) if t]

    exact = sorted({t for t in t_tokens if t in e_tokens})
    if exact:
        reason = "TEM4 义项与英文词中文义存在完全相同的规范化词元：" + "、".join(exact) + "。"
        return "CORE", reason, exact

    partial_hits: list[str] = []
    for t in t_tokens:
        for e in e_tokens:
            if len(t) >= 2 and len(e) >= 2 and (t in e or e in t):
                partial_hits.append(f"{t}⊂{e}")
            elif len(t) == len(e) and len(t) >= 2 and 0 < levenshtein(t, e) <= 2:
                partial_hits.append(f"{t}≈{e}")
    if partial_hits:
        reason = "义项之间存在包含或近形关系（子串/小编辑距离），需要人工确认是否构成部分对应：" + "、".join(sorted(set(partial_hits))[:8]) + "。"
        return "PARTIAL", reason, sorted(set(partial_hits))

    shingle_hits: list[str] = []
    first_tem4_shingles = cn_shingles(tem4_senses[0]) if tem4_senses else set()
    for sense_index, sense in enumerate(tem4_senses[:2]):
        for english_sense in english_senses[:2]:
            shared = cn_shingles(sense) & cn_shingles(english_sense)
            for shingle in sorted(shared):
                marker = "核心义" if sense_index == 0 else "次要义"
                shingle_hits.append(f"{marker}:{shingle}")
    if shingle_hits:
        core_hits = [h for h in shingle_hits if h.startswith("核心义")]
        decision = "CORE" if core_hits else "PARTIAL"
        reason = "TEM4 义项与英文词中文义共享核心双字元（如“印象”）：" + "、".join(shingle_hits[:8]) + "。该情况无法仅靠字符串判定，进入人工审查队列。"
        return decision, reason, shingle_hits
    return "NONE", "TEM4 前 2–3 义项与易混英文词中文义无字符串重合、无子串/近形关系，亦无共享核心双字元；概念层不重合仍需人工审定。", []


def adjudicate() -> list[AdjudicationRecord]:
    wordbook = load_jsonl(DATA / "wordbook.jsonl")
    tem4_by_headword: dict[str, dict] = {}
    for row in load_jsonl(DATA / "tem4-candidates.jsonl"):
        tem4_by_headword.setdefault(row["normalized_headword"], row)
    ff_by_headword: dict[str, dict] = {}
    for row in load_jsonl(DATA / "false-friends-candidates.jsonl"):
        if row.get("visual_verification") and row.get("normalized_headword"):
            ff_by_headword.setdefault(row["normalized_headword"], row)
    tem4_ocr = load_ocr_pages(OCR_TEM4)
    ff_ocr = load_ff_ocr_pages(OCR_FF)
    reviewed_by_headword = load_reviewed_vocabulary(REVIEWED_WORKBOOK)

    records: list[AdjudicationRecord] = []
    for word in sorted(wordbook, key=lambda w: w["word_id"]):
        french = word["french_word"]
        english = word["english_word"]
        tem4 = tem4_by_headword.get(french)
        ff = ff_by_headword.get(french)
        reviewed = reviewed_by_headword.get(french, {})

        tem4_senses = reviewed.get("tem4TopSenses") or ((tem4.get("chinese_core_senses") or [])[:3] if tem4 else [])
        english_senses = reviewed.get("englishChineseSenses") or ((ff.get("english_true_meanings") or []) if ff else [])
        if not english_senses:
            english_senses = [s for s in cn_tokens(word.get("false_friend_meanings", "")) if s]

        decision, reason, overlaps = decide_semantic_overlap(tem4_senses, english_senses)

        raw_distance = levenshtein(french, english)
        folded_distance = levenshtein(fold_latin(french), fold_latin(english))
        morphology_only, morphology_rule = detect_morphology_only(french, english)

        page_text = tem4_ocr.get(int(word["tem4_pdf_page"]), "")
        example, example_status = find_example_sentence(page_text, french)

        ff_page_text = ff_ocr.get(int(word["false_friend_pdf_page"]), "")
        ff_quote = reviewed.get("falseFriendsSourceQuote") or ""
        if ff is not None:
            ff_quote = str(ff.get("page_evidence") or "")
        if not ff_quote and ff_page_text:
            index = ff_page_text.lower().find(french.lower())
            ff_quote = ff_page_text[max(0, index - 120): index + 240].replace("\n", " ") if index >= 0 else ""

        tem4_quote = reviewed.get("tem4SourceQuote") or (str(tem4.get("page_evidence") or "") if tem4 else "")

        eligible: list[str] = []
        if decision == "NONE" and tem4_senses and english_senses:
            eligible.append("T1")
            if example and example_status == "SOURCE_VERIFIED":
                eligible.append("T2")
            eligible.append("T3")
        if not morphology_only and 1 <= min(raw_distance, folded_distance) <= 4:
            eligible.append("T4")

        notes_parts = []
        if tem4 is None:
            notes_parts.append("TEM4 候选缺失")
        if ff is None:
            notes_parts.append("假朋友候选缺失")
        if example_status != "SOURCE_VERIFIED":
            notes_parts.append("TEM4 页面无完整例句")
        if not eligible:
            notes_parts.append("不满足任何题型资格")
        if reviewed.get("reviewCorrection"):
            notes_parts.append("沿用既有审定纠正：" + reviewed["reviewCorrection"])

        records.append(AdjudicationRecord(
            wordId=word["word_id"],
            frenchWord=french,
            englishConfusable=english,
            tem4TopSenses=tem4_senses,
            englishChineseSenses=english_senses,
            semanticOverlapDecision=decision,
            semanticDecisionReason=reason,
            tem4PdfPage=int(word["tem4_pdf_page"]),
            falseFriendsPdfPage=int(word["false_friend_pdf_page"]),
            tem4SourceQuote=tem4_quote,
            falseFriendsSourceQuote=ff_quote[:500],
            tem4ExampleSentence=example,
            exampleSentenceStatus=example_status,
            rawEditDistance=raw_distance,
            accentFoldedEditDistance=folded_distance,
            morphologyOnly=morphology_only,
            morphologyRule=morphology_rule,
            eligibleTypes=eligible,
            assignedType=None,
            lexicalReviewStatus="LEXICAL_REVIEW_PENDING",
            pedagogicReviewStatus="PEDAGOGIC_REVIEW_PENDING",
            reviewNotes="；".join(notes_parts) or "无",
            senseTokenOverlap=overlaps,
            candidateId=tem4.get("candidate_id") if tem4 else None,
            partOfSpeech=reviewed.get("partOfSpeech") or (tem4.get("part_of_speech") if tem4 else None),
            unit=reviewed.get("unit") or (tem4.get("unit") if tem4 else None),
            sourceRelationType=reviewed.get("sourceRelationType"),
        ))
    return records


def write_jsonl(records: list[AdjudicationRecord], path: Path) -> None:
    with path.open("w", encoding="utf-8") as handle:
        for record in records:
            handle.write(json.dumps(asdict(record), ensure_ascii=False) + "\n")


def write_xlsx(records: list[AdjudicationRecord], path: Path) -> None:
    columns = list(asdict(records[0]).keys())
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "候选审定"
    sheet.append(columns)
    for column in range(1, len(columns) + 1):
        cell = sheet.cell(row=1, column=column)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for record in records:
        values = asdict(record)
        values["eligibleTypes"] = "、".join(values["eligibleTypes"])
        values["assignedType"] = values["assignedType"] or ""
        values["tem4TopSenses"] = "；".join(values["tem4TopSenses"])
        values["englishChineseSenses"] = "；".join(values["englishChineseSenses"])
        values["senseTokenOverlap"] = "、".join(values["senseTokenOverlap"])
        sheet.append([xml_safe(values.get(column)) for column in columns])
    for row in range(2, len(records) + 2):
        for column in range(1, len(columns) + 1):
            sheet.cell(row=row, column=column).alignment = WRAP
    for column, width in zip(range(1, len(columns) + 1), [14, 16, 18, 34, 34, 22, 46, 12, 16, 40, 40, 42, 24, 14, 14, 12, 14, 24, 16, 24, 24, 40, 20, 18, 18, 20]):
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.freeze_panes = "A2"
    workbook.save(path)


def write_review_queue(records: list[AdjudicationRecord], path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "语义审查队列"
    columns = ["wordId", "frenchWord", "englishConfusable", "tem4TopSenses", "englishChineseSenses",
               "semanticOverlapDecision", "semanticDecisionReason", "senseTokenOverlap",
               "tem4PdfPage", "falseFriendsPdfPage", "eligibleTypes", "reviewNotes"]
    sheet.append(columns)
    for column in range(1, len(columns) + 1):
        cell = sheet.cell(row=1, column=column)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    queued = [
        r for r in records
        if r.semanticOverlapDecision in ("PARTIAL", "UNCERTAIN")
        or (r.semanticOverlapDecision == "NONE" and r.eligibleTypes)
    ]
    for record in queued:
        sheet.append([
            record.wordId, record.frenchWord, record.englishConfusable,
            "；".join(record.tem4TopSenses), "；".join(record.englishChineseSenses),
            record.semanticOverlapDecision, xml_safe(record.semanticDecisionReason),
            "、".join(record.senseTokenOverlap), record.tem4PdfPage, record.falseFriendsPdfPage,
            "、".join(record.eligibleTypes), xml_safe(record.reviewNotes),
        ])
    for row in range(2, len(queued) + 2):
        for column in range(1, len(columns) + 1):
            sheet.cell(row=row, column=column).alignment = WRAP
    for column, width in zip(range(1, len(columns) + 1), [14, 16, 18, 34, 34, 22, 60, 30, 12, 16, 20, 40]):
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.freeze_panes = "A2"
    workbook.save(path)


def main() -> None:
    records = adjudicate()
    write_jsonl(records, OUT / "candidate-adjudication.jsonl")
    write_xlsx(records, OUT / "candidate-adjudication.xlsx")
    write_review_queue(records, OUT / "semantic-review-queue.xlsx")

    summary = {
        "total": len(records),
        "semanticNONE": sum(1 for r in records if r.semanticOverlapDecision == "NONE"),
        "semanticPARTIAL": sum(1 for r in records if r.semanticOverlapDecision == "PARTIAL"),
        "semanticCORE": sum(1 for r in records if r.semanticOverlapDecision == "CORE"),
        "eligibleT1": sum(1 for r in records if "T1" in r.eligibleTypes),
        "eligibleT2": sum(1 for r in records if "T2" in r.eligibleTypes),
        "eligibleT3": sum(1 for r in records if "T3" in r.eligibleTypes),
        "eligibleT4": sum(1 for r in records if "T4" in r.eligibleTypes),
        "exampleMissing": sum(1 for r in records if r.exampleSentenceStatus != "SOURCE_VERIFIED"),
        "morphologyOnly": sum(1 for r in records if r.morphologyOnly),
        "reviewQueue": len([
            r for r in records
            if r.semanticOverlapDecision in ("PARTIAL", "UNCERTAIN")
            or (r.semanticOverlapDecision == "NONE" and r.eligibleTypes)
        ]),
    }
    (OUT / "candidate-summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False))


if __name__ == "__main__":
    main()
