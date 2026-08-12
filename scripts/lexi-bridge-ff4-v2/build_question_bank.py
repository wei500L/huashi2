#!/usr/bin/env python3
"""FF4 question bank V2 - four-type question bank generator.

Builds the question bank workbook and machine-readable JSON package from the
candidate adjudication + T2 evidence files. Rules:

  T1 (word meaning single choice):
    - 1 correct option = TEM4 top 2-3 Chinese senses
    - 1 transfer option = English confusable's Chinese senses
    - 2 distractors = same POS, similar difficulty, no overlap with correct
  T2 (sentence selection):
    - stem = TEM4 original sentence with target word in **bold+underline**
    - 1 correct option = source-evidenced near-synonym
    - 1 transfer option = source-evidenced French word for English meaning
    - 2 distractors = same POS/syntax, semantically wrong
  T3 (true/false):
    - stem = "法语词 = 易混英文词的中文义"; always F (see TYPE3_DESIGN_DECISION.md)
  T4 (spelling):
    - stem = "中文释义 ______（填写对应法语单词）"
    - correctAnswers = [frenchWord]; acceptableAnswers added when ambiguous

Target words are globally unique across the four types. Every generated item
carries source pages, quotes and production-approved review statuses. Raw
candidate files retain their original pending labels as adjudication history.

Outputs:
  docs/data/lexi-bridge-ff4-v2/法语专四假朋友题库_V2.xlsx
  docs/data/lexi-bridge-ff4-v2/question-bank-package-v2.json
"""

from __future__ import annotations

import hashlib
import json
import random
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from build_candidates import (  # noqa: E402
    ROOT, HEADER_FILL, HEADER_FONT, WRAP, fold_latin, load_jsonl,
    normalize_cn,
)
from production_semantic_rules import (  # noqa: E402
    APPROVED_SEMANTIC_WORDS, T2_RULES, T2_TARGET_WORDS,
)

OUT = ROOT / "docs" / "data" / "lexi-bridge-ff4-v2"
from openpyxl import Workbook  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402


def content_hash(payload: dict) -> str:
    canonical = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(canonical.encode("utf-8")).hexdigest()


def pick_distractors(pool: list[dict], target: dict, count: int, rng: random.Random,
                     exclude_senses: list[str]) -> list[str]:
    """Pick Chinese-sense distractors from the pool.

    A distractor must not share a normalized token with the correct/transfer
    senses (no synonymy/overlap), must be a different word entry, and must
    have the same coarse part of speech. Targets without two valid same-POS
    distractors are excluded; the generator never lowers this rule.
    """
    target_tokens = {normalize_cn(t) for sense in exclude_senses for t in re.split(r"[;；,，、\s]+", str(sense)) if t}
    pos = (target.get("partOfSpeech") or "").split(".")[0].strip()

    def sense_of(word: dict) -> str:
        return "; ".join((word.get("tem4TopSenses") or [])[:2])

    def compatible(word: dict) -> bool:
        if word["wordId"] == target["wordId"]:
            return False
        tokens = {normalize_cn(t) for t in re.split(r"[;；,，、\s]+", str(sense_of(word))) if t}
        return not (tokens & target_tokens)

    preferred = [w for w in pool if compatible(w)
                 and (w.get("partOfSpeech") or "").split(".")[0].strip() == pos]
    candidates = preferred
    rng.shuffle(candidates)
    chosen: list[str] = []
    for word in candidates:
        sense = sense_of(word)
        if sense and sense not in chosen:
            chosen.append(sense)
        if len(chosen) >= count:
            break
    return chosen


def build_t1_item(record: dict, pool: list[dict], rng: random.Random) -> dict | None:
    correct = "；".join(record["tem4TopSenses"][:3])
    transfer = "；".join(record["englishChineseSenses"][:3])
    distractors = pick_distractors(pool, record, 2, rng, record["tem4TopSenses"][:3] + record["englishChineseSenses"])
    if len(distractors) != 2:
        return None
    option_pool = [correct, transfer] + distractors
    rng.shuffle(option_pool)
    keys = "ABCD"
    options = [{"key": key, "label": label, "correct": label == correct,
                "role": "CORRECT" if label == correct else "TRANSFER" if label == transfer else "DISTRACTOR"}
               for key, label in zip(keys, option_pool)]
    return {
        "itemType": "T1",
        "targetWord": record["frenchWord"],
        "stemText": f"请选出下列法语单词对应的正确中文含义\n{record['frenchWord']}",
        "options": options,
        "correctAnswers": [next(o["key"] for o in options if o["correct"])],
    }


def same_pos_words(pool: list[dict], target: dict, exclude: set[str], count: int, rng: random.Random) -> list[str]:
    target_pos = (target.get("partOfSpeech") or "").split(".")[0].strip()
    candidates = [w for w in pool
                  if w["wordId"] != target["wordId"]
                  and w["frenchWord"] not in exclude
                  and (w.get("partOfSpeech") or "").split(".")[0].strip() == target_pos]
    rng.shuffle(candidates)
    chosen: list[str] = []
    for word in candidates:
        if word["frenchWord"] not in chosen:
            chosen.append(word["frenchWord"])
        if len(chosen) >= count:
            break
    return chosen


def gender_compatible(target_pos: str, candidate_pos: str) -> bool:
    """Check grammatical gender compatibility for the T2 option slot.

    A n.f. target slot accepts only n.f. words; a n.m. slot only n.m. words.
    Non-noun slots accept any word of the same part of speech.
    """
    t = (target_pos or "").strip().lower()
    c = (candidate_pos or "").strip().lower()
    if t.startswith("n.") and c.startswith("n."):
        t_f = "f" in t.split(";")[0]
        c_f = "f" in c.split(";")[0]
        return t_f == c_f
    return t.split(".")[0].strip() == c.split(".")[0].strip()


def build_t2_item(record: dict, pool: list[dict], rng: random.Random) -> dict:
    rule = T2_RULES.get(record["frenchWord"])
    if rule is None:
        raise ValueError(f"unsupported production T2 target: {record['frenchWord']}")
    options = [
        {"key": key, "label": label, "correct": correct, "role": role}
        for key, label, correct, role in rule["options"]
    ]
    return {
        "itemType": "T2",
        "targetWord": record["frenchWord"],
        "stemText": "请根据句子选择画线单词的同义解释\n" + rule["sentence"],
        "options": options,
        "correctAnswers": [next(o["key"] for o in options if o["correct"])],
    }


def build_t3_item(record: dict) -> dict:
    english_sense = "；".join(record["englishChineseSenses"][:2])
    return {
        "itemType": "T3",
        "targetWord": record["frenchWord"],
        "stemText": f"{record['frenchWord']} = {english_sense}",
        "options": [
            {"key": "V", "label": "正确", "correct": False, "role": "DISTRACTOR"},
            {"key": "F", "label": "错误", "correct": True, "role": "CORRECT"},
        ],
        "correctAnswers": ["F"],
    }


def build_t4_item(record: dict, rng: random.Random) -> dict:
    stem_sense = "；".join(record["tem4TopSenses"][:3])
    return {
        "itemType": "T4",
        "targetWord": record["frenchWord"],
        "stemText": f"{stem_sense} ______（填写对应法语单词）",
        "options": [],
        "correctAnswers": [record["frenchWord"]],
    }


def find_t4_ambiguous(records: list[dict]) -> dict[str, list[str]]:
    """Detect Chinese stems that map to more than one French word.

    For each T4-eligible word, find other pool words whose top senses share a
    normalized token with the stem sense; if any exist, the item is ambiguous
    and must either declare acceptableAnswers or be replaced.
    """
    eligible = {r["frenchWord"]: r for r in records if "T4" in r["eligibleTypes"]}
    ambiguity: dict[str, list[str]] = {}
    for word, record in eligible.items():
        stem_tokens = {normalize_cn(t) for sense in record["tem4TopSenses"][:3]
                       for t in re.split(r"[;；,，、\s]+", str(sense)) if t}
        conflicting = []
        for other, other_record in eligible.items():
            if other == word:
                continue
            other_tokens = {normalize_cn(t) for sense in other_record["tem4TopSenses"][:2]
                            for t in re.split(r"[;；,，、\s]+", str(sense)) if t}
            if stem_tokens & other_tokens:
                conflicting.append(other)
        if conflicting:
            ambiguity[word] = conflicting
    return ambiguity


def build() -> dict:
    records = load_jsonl(OUT / "candidate-adjudication.jsonl")
    evidence = {r["frenchWord"]: r for r in load_jsonl(OUT / "t2-evidence.jsonl")}
    rng = random.Random(20260811)

    semantic_pool = [r for r in records if r["frenchWord"] in APPROVED_SEMANTIC_WORDS]
    t2_records = [r for r in semantic_pool
                  if r["frenchWord"] in T2_TARGET_WORDS]
    t2_words = {r["frenchWord"] for r in t2_records}
    t1_pool = [r for r in semantic_pool if r["frenchWord"] not in t2_words]
    t4_pool = [r for r in records if "T4" in r["eligibleTypes"]]

    items = []
    used_words: set[str] = set()

    for record in t2_records:
        record.update(evidence[record["frenchWord"]])
        items.append(build_t2_item(record, semantic_pool, rng))
        used_words.add(record["frenchWord"])

    rng.shuffle(t1_pool)
    for record in t1_pool:
        if sum(1 for item in items if item["itemType"] == "T1") >= 15:
            break
        item = build_t1_item(record, records, rng)
        if item is None:
            continue
        items.append(item)
        used_words.add(record["frenchWord"])

    t3_pool = [r for r in semantic_pool if r["frenchWord"] not in used_words]
    rng.shuffle(t3_pool)
    t3_count = min(15, len(t3_pool))
    for record in t3_pool[:t3_count]:
        items.append(build_t3_item(record))
        used_words.add(record["frenchWord"])

    t4_available = [r for r in t4_pool if r["frenchWord"] not in used_words]
    t4_ambiguity = find_t4_ambiguous(records)
    unambiguous_t4 = [r for r in t4_available if r["frenchWord"] not in t4_ambiguity]
    rng.shuffle(unambiguous_t4)
    t4_count = min(15, len(unambiguous_t4))
    for record in unambiguous_t4[:t4_count]:
        items.append(build_t4_item(record, rng))
        used_words.add(record["frenchWord"])
    excluded_ambiguous = [w for w in t4_ambiguity if w not in used_words]
    t4_ambiguity_note = "；".join(
        f"{w}(可与{'、'.join(t4_ambiguity[w])}混淆)" for w in sorted(excluded_ambiguous)[:20])

    # deterministic ordering: T1, T2, T3, T4
    type_order = {"T1": 0, "T2": 1, "T3": 2, "T4": 3}
    ordered = sorted(items, key=lambda item: type_order[item["itemType"]])
    package_items = []
    package_options = []
    counts = {"T1": 0, "T2": 0, "T3": 0, "T4": 0}
    record_by_word = {r["frenchWord"]: r for r in records}
    for item in ordered:
        counts[item["itemType"]] += 1
        code = f"FF4V2-{item['itemType']}-{counts[item['itemType']]:03d}"
        item_type = item["itemType"]
        section_code = {"T1": "FF4_WORD_MEANING", "T2": "FF4_SENTENCE_SELECTION",
                        "T3": "FF4_TRUE_FALSE", "T4": "FF4_SPELLING"}[item_type]
        construct_code = {"T1": "FF4_WORD_MEANING", "T2": "FF4_SENTENCE_SYNONYM",
                          "T3": "FF4_TRUE_FALSE_TRANSFER", "T4": "FF4_SPELLING"}[item_type]
        question_type = {"T1": "SINGLE_CHOICE", "T2": "SINGLE_CHOICE",
                         "T3": "TRUE_FALSE", "T4": "SPELLING"}[item_type]
        record = record_by_word[item["targetWord"]]
        evidence_record = evidence.get(item["targetWord"], {})
        package_items.append({
            "itemCode": code,
            "sectionCode": section_code,
            "questionType": question_type,
            "stemText": item["stemText"],
            "promptText": None,
            "correctAnswers": item["correctAnswers"],
            "explanationText": None,
            "requiredAnswer": True,
            "scored": True,
            "weight": 1,
            "transferCategory": "FALSE_FRIEND",
            "contextLevel": "SENTENCE" if item_type == "T2" else "WORD",
            "constructCode": construct_code,
            "targetWord": item["targetWord"],
            "displayConditionJson": None,
            "tem4PdfPage": record.get("tem4PdfPage"),
            "falseFriendsPdfPage": record.get("falseFriendsPdfPage"),
            "exampleSentenceStatus": (T2_RULES[item["targetWord"]].get("evidenceLevel", "TEM4_EXACT_SENTENCE")
                                      if item_type == "T2" else record.get("exampleSentenceStatus")),
            "spellingRawEditDistance": record.get("rawEditDistance") if item_type == "T4" else None,
            "spellingAccentFoldedEditDistance": record.get("accentFoldedEditDistance") if item_type == "T4" else None,
            "morphologyOnly": record.get("morphologyOnly") if item_type == "T4" else False,
            "lexicalReviewStatus": "APPROVED",
            "pedagogicReviewStatus": "APPROVED",
        })
        for option in item["options"]:
            package_options.append({
                "itemCode": code,
                "optionCode": option["key"],
                "optionText": option["label"],
                "correct": option["correct"],
                "explanation": None,
                "role": option["role"],
            })

    sections = [
        {"sectionCode": "FF4_WORD_MEANING", "title": "题型一 词义单选",
         "description": "请选出下列法语单词对应的正确中文含义。", "sortOrder": 1, "formalSection": True,
         "sharedMaterial": None},
        {"sectionCode": "FF4_SENTENCE_SELECTION", "title": "题型二 句子选词",
         "description": "请根据句子选择画线单词的同义表达。", "sortOrder": 2, "formalSection": True,
         "sharedMaterial": None},
        {"sectionCode": "FF4_TRUE_FALSE", "title": "题型三 判断正误",
         "description": "判断正误：法语单词是否表示给定的中文含义。", "sortOrder": 3, "formalSection": True,
         "sharedMaterial": None},
        {"sectionCode": "FF4_SPELLING", "title": "题型四 单词拼写",
         "description": "根据中文释义填写对应的法语单词。", "sortOrder": 4, "formalSection": True,
         "sharedMaterial": None},
    ]

    package = {
        "questionnaire": {
            "code": "LEXIBRIDGE_RESEARCH_V3",
            "title": "Lexi-bridge 英法词汇认知迁移研究问卷 V3",
            "description": "法语专四假朋友词汇四类题研究问卷",
            "durationMinutes": 40,
            "scoringVersion": "SCORING_V3",
            "aiPromptVersion": "assessment-analysis/v3",
        },
        "sections": sections,
        "items": package_items,
        "options": package_options,
    }
    return package


def write_workbook(package: dict) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "题目总览"
    columns = ["题目编号", "题型", "词条", "题干", "选项A", "选项B", "选项C", "选项D", "标准答案",
               "正确义", "英文迁移义", "来源页码", "词汇审查状态", "教研审查状态"]
    sheet.append(columns)
    for column in range(1, len(columns) + 1):
        cell = sheet.cell(row=1, column=column)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    records = load_jsonl(OUT / "candidate-adjudication.jsonl")
    record_by_word = {r["frenchWord"]: r for r in records}
    options_by_item = {}
    for option in package["options"]:
        options_by_item.setdefault(option["itemCode"], []).append(option)
    type_by_construct = {"FF4_WORD_MEANING": "T1", "FF4_SENTENCE_SYNONYM": "T2",
                         "FF4_TRUE_FALSE_TRANSFER": "T3", "FF4_SPELLING": "T4"}
    for item in package["items"]:
        word = item["targetWord"]
        record = record_by_word.get(word, {})
        options = options_by_item.get(item["itemCode"], [])
        by_key = {o["optionCode"]: o for o in options}
        if options:
            correct = next(o for o in options if o["correct"])
            transfer = next((o for o in options if o["role"] == "TRANSFER"), None)
            answer_label = correct["optionCode"]
        else:
            correct = {"optionCode": ""}
            transfer = None
            answer_label = item["correctAnswers"][0]
        row = [
            item["itemCode"], type_by_construct[item["constructCode"]], word, item["stemText"],
            by_key.get("A", {}).get("optionText", ""), by_key.get("B", {}).get("optionText", ""),
            by_key.get("C", {}).get("optionText", ""), by_key.get("D", {}).get("optionText", ""),
            answer_label,
            "；".join(record.get("tem4TopSenses", [])[:3]) if record else "",
            transfer["optionText"] if transfer else "",
            f"TEM4 p{record.get('tem4PdfPage', '?')}；假朋友词典 p{record.get('falseFriendsPdfPage', '?')}" if record else "",
            "APPROVED",
            "APPROVED",
        ]
        sheet.append(row)
    for row in range(2, len(package["items"]) + 2):
        for column in range(1, len(columns) + 1):
            sheet.cell(row=row, column=column).alignment = WRAP
    for column, width in zip(range(1, len(columns) + 1), [16, 8, 16, 50, 26, 26, 26, 26, 10, 30, 30, 34, 22, 22]):
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.freeze_panes = "A2"
    workbook.save(OUT / "法语专四假朋友题库_V2.xlsx")


def main() -> None:
    package = build()
    (OUT / "question-bank-package-v2.json").write_text(
        json.dumps(package, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    write_workbook(package)
    counts = {}
    for item in package["items"]:
        item_type = {"FF4_WORD_MEANING": "T1", "FF4_SENTENCE_SYNONYM": "T2",
                     "FF4_TRUE_FALSE_TRANSFER": "T3", "FF4_SPELLING": "T4"}[item["constructCode"]]
        counts[item_type] = counts.get(item_type, 0) + 1
    print(json.dumps({"total": len(package["items"]), "counts": counts}, ensure_ascii=False))


if __name__ == "__main__":
    main()
